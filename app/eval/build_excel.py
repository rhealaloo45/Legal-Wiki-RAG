"""Export the top-90-by-score subset of the audit to Excel.

The workbook leads with a Selection sheet because the export is a filtered view:
dropping the 15 lowest-scoring questions raises the mean by roughly a point, so
the subset's average is not the system's score and the file has to say which
number is which without the reader having to reconstruct it.
"""

import json
import os
import sys

sys.path.insert(0, os.path.dirname(__file__))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from build_artifact import INGEST_CAPPED, RETEST, V1

TEMP = os.environ.get("TEMP", ".")
OUT = os.path.join(os.path.dirname(__file__), "..", "..", "rag_top90_review.xlsx")
KEEP = 90

INK = "1F2A2D"
ACCENT = "2C5F66"
HEAD_FILL = PatternFill("solid", fgColor=ACCENT)
BANDS = {"pass": PatternFill("solid", fgColor="E7F0EA"),
         "partial": PatternFill("solid", fgColor="F6EFDF"),
         "fail": PatternFill("solid", fgColor="F7E7E5")}
THIN = Side(style="thin", color="D7E0E1")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TOP_WRAP = Alignment(vertical="top", wrap_text=True)


def band(s: int) -> str:
    return "pass" if s >= 8 else ("partial" if s >= 5 else "fail")


def main() -> None:
    data = json.load(open(os.path.join(TEMP, "full_report_data.json"), encoding="utf-8"))
    gap = os.path.join(TEMP, "gap_out.json")
    if os.path.exists(gap):
        for r in json.load(open(gap, encoding="utf-8")):
            if r["id"] in data and r.get("answer"):
                data[r["id"]]["answer"] = r["answer"]

    rows = []
    for n in range(1, 106):
        e = data.get(f"Q{n}", {})
        rows.append({
            "n": n, "q": e.get("question", ""), "gt": e.get("gt", ""),
            "src": e.get("src", ""), "answer": (e.get("answer") or "").strip(),
            "score": RETEST.get(n, V1[n]),
        })

    full_mean = sum(r["score"] for r in rows) / len(rows)
    # Ties at the cut are broken by question number so the selection is reproducible.
    kept = sorted(sorted(rows, key=lambda r: (-r["score"], r["n"]))[:KEEP],
                  key=lambda r: r["n"])
    dropped = sorted(set(r["n"] for r in rows) - set(r["n"] for r in kept))
    kept_mean = sum(r["score"] for r in kept) / len(kept)
    cutoff = min(r["score"] for r in kept)

    wb = Workbook()

    s = wb.active
    s.title = "Selection"
    s.column_dimensions["A"].width = 34
    s.column_dimensions["B"].width = 92
    notes = [
        ("LexWiki RAG accuracy audit", ""),
        ("", ""),
        ("What this file contains", f"The {KEEP} highest-scoring of the 105 audited questions."),
        ("What it excludes", f"The {len(dropped)} lowest-scoring questions: "
                             + ", ".join(f"Q{n}" for n in dropped)),
        ("Lowest score included", f"{cutoff}/10"),
        ("", ""),
        ("Mean of this subset", f"{kept_mean:.2f} / 10"),
        ("Mean of all 105 questions", f"{full_mean:.2f} / 10  <- the system's score"),
        ("Prior audit mean (all 105)", f"{sum(V1.values()) / len(V1):.2f} / 10"),
        ("", ""),
        ("Read this before quoting a number",
         f"This is a filtered subset, so {kept_mean:.2f} is not the system's accuracy. "
         f"It is the average over the {KEEP} questions it answered best, with the "
         f"{len(dropped)} weakest removed. The comparable figure is {full_mean:.2f}."),
        ("", ""),
        ("Scoring basis", "Manual scoring against verified ground truth. Re-tested questions "
                          "show the better of two corpus runs (all 494 documents, or the 46 "
                          "real documents with synthetic Test_* fixtures removed); the rest "
                          "retain their original mixed-corpus score."),
        ("Ingest-capped questions", "For 8 questions the ground-truth fact is absent from every "
                                    "ingested page of the expected document, so the answer text "
                                    "is not in the database at all. Those refusals are correct "
                                    "and no retrieval change can raise them."),
    ]
    for i, (k, v) in enumerate(notes, start=1):
        s.cell(i, 1, k).font = Font(bold=bool(k), color=INK, size=13 if i == 1 else 11)
        c = s.cell(i, 2, v)
        c.alignment = TOP_WRAP
        if k.startswith("Read this"):
            c.font = Font(bold=True, color="9A3A31")
    s.row_dimensions[11].height = 46
    s.row_dimensions[13].height = 46
    s.row_dimensions[14].height = 46

    d = wb.create_sheet("Top 90 questions")
    headers = ["#", "Question", "Application answer", "Grounded answer",
               "Source document", "Score", "Notes"]
    widths = [6, 46, 68, 58, 38, 8, 16]
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        c = d.cell(1, i, h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = HEAD_FILL
        c.alignment = Alignment(vertical="center")
        c.border = BORDER
        d.column_dimensions[get_column_letter(i)].width = w
    d.freeze_panes = "B2"
    d.auto_filter.ref = f"A1:G{len(kept) + 1}"

    for i, r in enumerate(kept, start=2):
        note = "ingest-capped" if r["n"] in INGEST_CAPPED else (
            "re-tested" if r["n"] in RETEST else "")
        for col, val in enumerate([f"Q{r['n']}", r["q"], r["answer"], r["gt"],
                                   r["src"], r["score"], note], start=1):
            c = d.cell(i, col, val)
            c.alignment = TOP_WRAP if col in (2, 3, 4, 5) else Alignment(vertical="top")
            c.border = BORDER
            if col == 6:
                c.fill = BANDS[band(r["score"])]
                c.font = Font(bold=True)
                c.alignment = Alignment(vertical="top", horizontal="center")
        d.row_dimensions[i].height = 96

    out = os.path.abspath(OUT)
    wb.save(out)
    print("wrote", out)
    print(f"kept {len(kept)} (mean {kept_mean:.2f}, cutoff {cutoff}) | "
          f"dropped {len(dropped)}: {', '.join('Q%d' % n for n in dropped)}")
    print(f"full-set mean {full_mean:.2f}")


if __name__ == "__main__":
    main()
