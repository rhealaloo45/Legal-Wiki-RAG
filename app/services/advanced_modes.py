import os
import io
import re
import json
import uuid
import logging
import concurrent.futures
from threading import Lock
from typing import Optional

import config
from services import reader, llm, wiki

logger = logging.getLogger(__name__)

# Max documents processed in a single Review/Compare background job pass.
# Larger selections are chunked into multiple ≤_MAX_DOCS_PER_RUN passes within
# the same job (same job_id, same progress store) rather than firing one
# unbounded batch of concurrent LLM calls.
_MAX_DOCS_PER_RUN = 20


def _chunk_docs(docs: list[str]) -> list[list[str]]:
    return [docs[i:i + _MAX_DOCS_PER_RUN] for i in range(0, len(docs), _MAX_DOCS_PER_RUN)]


# Shared by Review and Compare: mapping from normalised column/aspect names to
# page_metadata field names. When a requested column matches a standard field
# that was pre-extracted at ingest time, serve it directly from the DB —
# no LLM call needed.
_METADATA_FIELD_MAP = {
    "governing law": "governing_law",
    "jurisdiction": "jurisdiction",
    "effective date": "effective_date",
    "termination notice": "termination_notice",
    "termination notice period": "termination_notice",
    "liability cap": "liability_cap",
    "ip ownership": "ip_ownership",
    "intellectual property ownership": "ip_ownership",
    "parties": "parties",
    "auto renewal": "auto_renewal",
    "auto-renewal": "auto_renewal",
    "notice period": "notice_period",
    "payment terms": "payment_terms",
}

# Column/aspect names that read as open-ended asks (summaries, overviews) —
# these need a wider slice of the document up front rather than waiting for a
# low-confidence retry, since a narrow retrieval slice is unlikely to cover
# "summarize this agreement".
_OPEN_ENDED_COLUMN_RE = re.compile(
    r"\b(summar|overview|key\s*terms|description|explain|analysis|synopsis)", re.IGNORECASE
)


def _is_open_ended_column(col_name: str) -> bool:
    return bool(_OPEN_ENDED_COLUMN_RE.search(col_name))


def _lookup_cached_metadata(session_id: str, doc_name: str, col_name: str) -> Optional[dict]:
    """Try the page_metadata cache before firing an LLM call. Returns a
    ready-to-store result dict, or None if there's no cache hit."""
    field_key = _METADATA_FIELD_MAP.get(col_name.lower().strip())
    if not (field_key and config.USE_DATABASE):
        return None
    try:
        from services import db as _db, wikis as _wikis
        cached = _db.get_metadata(_wikis.active_wiki_id(), session_id, doc_name)
        if cached.get(field_key) is not None:
            return {"value": cached[field_key], "confidence": 0.95, "quote": None}
    except Exception as _e:
        logger.warning(f"Metadata cache lookup failed for {doc_name}/{col_name}: {_e}")
    return None

# ---------------------------------------------------------------------------
# Shared Utilities
# ---------------------------------------------------------------------------

def get_raw_doc_text(session_id: str, doc_name: str) -> str:
    """Retrieve raw text from an uploaded document."""
    flat_name = doc_name.replace("/", "_").replace("\\", "_")
    file_path = os.path.join(config.UPLOAD_PATH, f"{session_id}_{flat_name}")
    
    if not os.path.exists(file_path):
        # Resilient fallback: search across all uploads for any session matching flat_name
        found = False
        for fname in os.listdir(config.UPLOAD_PATH):
            if fname.endswith(flat_name):
                file_path = os.path.join(config.UPLOAD_PATH, fname)
                found = True
                break
        if not found:
            logger.error(f"Failed to find raw doc text at path: {file_path} or any fallback")
            return ""
            
    try:
        return reader.read_file(file_path)
    except Exception as e:
        logger.error(f"Error reading file {file_path}: {e}")
        return ""

# Context-char budgets for a single extraction cell. Narrow is the default —
# cheap, and sufficient for a targeted factual column/aspect. Broad is a
# capped escalation (same "cheap default, capped widen" shape as
# config.BROAD_QUESTION_TOTAL_CAP in wiki.py), used only for open-ended/summary
# columns or when a narrow-budget first pass comes back low-confidence/empty.
_CELL_CONTEXT_NARROW_LIMIT = 3000
_CELL_CONTEXT_BROAD_LIMIT = 8000


def _get_wiki_text_for_doc(session_id: str, doc_name: str, query: str = "", broad: bool = False) -> str:
    """Retrieve synthesized wiki content for a document, scoped to `query`.

    Looks up wiki pages tagged with this source_doc, then — instead of
    concatenating every one of them and truncating at a flat char limit —
    ranks them with the same vector+BM25 hybrid retrieval the main chat
    pipeline uses (wiki._select_relevant_pages), restricted to just this
    document's own pages, and fills a char budget with only the pages that
    actually rank for `query`. Falls back to raw text if no wiki pages are
    found, and to the old dump-everything-then-truncate behaviour if page
    selection fails or no query is given.

    broad=True widens the char budget for an open-ended column/aspect or an
    escalation retry after a narrow-budget pass came back low-confidence.

    The output preserves source attribution and supporting quotes embedded
    during wiki ingest, so downstream LLM calls can cite them.
    """
    limit = _CELL_CONTEXT_BROAD_LIMIT if broad else _CELL_CONTEXT_NARROW_LIMIT
    scoped = _get_scoped_wiki_pages(session_id, doc_name)
    if scoped:
        ordered_titles = list(scoped.keys())
        if query and config.USE_DATABASE:
            try:
                selected_titles, _usage = wiki._select_relevant_pages(
                    scoped, query, session_id=session_id, force_broad=broad,
                )
                ranked = [t for t in selected_titles if t in scoped]
                if ranked:
                    ordered_titles = ranked
            except Exception as e:
                logger.warning(
                    f"Scoped page selection failed for {doc_name}/{query[:40]!r}: {e} — using all scoped pages"
                )

        parts = [f"[Source Document: {doc_name}]"]
        total_len = 0
        for title in ordered_titles:
            page_data = scoped[title]
            content = page_data.get("content", "") if isinstance(page_data, dict) else str(page_data)
            summary = page_data.get("summary", "") if isinstance(page_data, dict) else ""
            chunk = f"## {title}\n{summary}\n{content}"
            if total_len and total_len + len(chunk) > limit:
                break
            parts.append(chunk)
            total_len += len(chunk)
        wiki_text = "\n\n".join(parts)
        if len(wiki_text) > 200:  # Only use wiki if substantial content exists
            return wiki_text

    # Fallback to raw text — prefix with source doc name for traceability
    raw = get_raw_doc_text(session_id, doc_name)
    if raw:
        return f"[Source Document: {doc_name}]\n\n{raw[:limit]}"
    return ""

def extract_cell(doc_text: str, column_name: str) -> dict:
    """Extract a specific piece of information from text using fast LLM path."""
    prompt = f"""\
Extract the specific piece of information requested from the legal text below.
Return JSON only, no preamble:
{{"value": str|null, "confidence": float 0-1, "quote": str|null}}

STRICT RULES:
- The "value" field must be a concise, informative summary (1-3 sentences) explaining the extracted information clearly, not just a single keyword.
- ONLY extract information that is EXPLICITLY stated in the text below.
- If the requested information is NOT found in the text, you MUST return null for the "value" field. Do NOT write sentences explaining that it is missing.
- DO NOT infer, assume, or hallucinate any values.
- The "quote" field MUST be a verbatim excerpt copied exactly from the text (max 120 chars).
  It serves as evidence for your extracted value.
- If the text contains "Supporting Quotes" sections, prefer those as your quote source.
- If the information is not found anywhere in the text: value=null, confidence=0.0, quote=null.
- confidence should reflect how directly the text states the information:
  1.0 = exact verbatim match, 0.8 = clearly stated, 0.5 = implied, 0.0 = not found.

Text:
{doc_text[:_CELL_CONTEXT_BROAD_LIMIT]}

Extract: {column_name}"""

    try:
        raw, _ = llm.fast_ask(prompt, max_tokens=300)
        # remove potential reasoning block or markdown
        raw = re.sub(r'<reasoning>.*?</reasoning>', '', raw, flags=re.DOTALL)
        raw = re.sub(r'```json', '', raw)
        raw = re.sub(r'```', '', raw)
        parsed = json.loads(raw.strip())
        return {
            "value": parsed.get("value"),
            "confidence": float(parsed.get("confidence", 0.0)),
            "quote": parsed.get("quote")
        }
    except Exception as e:
        logger.error(f"Cell extraction failed for '{column_name}': {e}")
        return {"value": None, "confidence": 0.0, "quote": None}


def _extract_with_retrieval(session_id: str, doc_name: str, query_text: str, broad_hint: bool = False) -> dict:
    """Fetch document-scoped context and extract a value, escalating once to
    the broader context budget if a narrow-budget first pass comes back
    low-confidence or empty. Shared by Review and Compare cell workers.
    """
    doc_text = _get_wiki_text_for_doc(session_id, doc_name, query=query_text, broad=broad_hint)
    res = extract_cell(doc_text, query_text)
    if not broad_hint and (res.get("confidence", 0.0) < 0.5 or res.get("value") is None):
        broad_text = _get_wiki_text_for_doc(session_id, doc_name, query=query_text, broad=True)
        if broad_text and broad_text != doc_text:
            res_broad = extract_cell(broad_text, query_text)
            if res_broad.get("value") is not None and res_broad.get("confidence", 0.0) >= res.get("confidence", 0.0):
                res = res_broad
    return res


def _apply_cell_styling(ws, cell, confidence):
    """Apply color styling based on confidence."""
    from openpyxl.styles import PatternFill
    if confidence is None:
        confidence = 0.0
        
    if confidence >= 0.8:
        fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") # green
    elif confidence >= 0.5:
        fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid") # yellow
    else:
        fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") # red
    
    cell.fill = fill

def _build_review_sheet(ws, store_data: dict):
    """docs=rows, columns=cols"""
    from openpyxl.styles import Font
    columns = store_data.get("columns", [])
    rows = store_data.get("rows", {})
    
    # Header
    ws.append(["Document Name"] + columns)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"
    
    # Rows
    for r_idx, (doc_name, col_data) in enumerate(rows.items(), start=2):
        row_values = [doc_name]
        for col_name in columns:
            cell_data = col_data.get(col_name, {})
            row_values.append(cell_data.get("value"))
            
        ws.append(row_values)
        
        # Style
        for c_idx, col_name in enumerate(columns, start=2):
            cell_data = col_data.get(col_name, {})
            confidence = cell_data.get("confidence", 0.0)
            _apply_cell_styling(ws, ws.cell(row=r_idx, column=c_idx), confidence)

def _build_compare_sheet(ws, store_data: dict):
    """aspects=rows, docs=cols"""
    from openpyxl.styles import Font
    sources = store_data.get("sources", [])
    aspects = store_data.get("aspects", [])
    table = store_data.get("table", {})
    
    doc_headers = [s.get("label", s.get("name")) for s in sources]
    
    # Header
    ws.append(["Aspect"] + doc_headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"
    
    # Rows
    for r_idx, aspect in enumerate(aspects, start=2):
        row_values = [aspect]
        aspect_data = table.get(aspect, {})
        for s in sources:
            doc_key = s.get("label", s.get("name"))
            cell_data = aspect_data.get(doc_key, {})
            row_values.append(cell_data.get("value"))
            
        ws.append(row_values)
        
        # Style
        for c_idx, s in enumerate(sources, start=2):
            doc_key = s.get("label", s.get("name"))
            cell_data = aspect_data.get(doc_key, {})
            confidence = cell_data.get("confidence", 0.0)
            _apply_cell_styling(ws, ws.cell(row=r_idx, column=c_idx), confidence)

def export_matrix_to_xlsx(store_data: dict, mode: str) -> bytes:
    """Generate Excel bytes for either review or compare store data."""
    try:
        import openpyxl
    except ImportError:
        logger.error("openpyxl not installed.")
        return b""
        
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Export"
    
    if mode == "review":
        _build_review_sheet(ws, store_data)
    elif mode == "compare":
        _build_compare_sheet(ws, store_data)
        
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter
    
    # Auto-adjust column widths and enable text wrapping for all cells
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 35 if col_idx == 1 else 55
        
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()

# ---------------------------------------------------------------------------
# Background Jobs
# ---------------------------------------------------------------------------

# Timeout for individual future results (seconds)
_FUTURE_TIMEOUT = 90

def _get_session_file_paths(session_id: str) -> list[str]:
    import json
    file_paths = []
    if os.path.exists(config.SESSIONS_PATH):
        try:
            with open(config.SESSIONS_PATH, "r", encoding="utf-8") as f:
                sessions = json.load(f)
                file_paths = sessions.get(session_id, {}).get("file_paths", [])
        except Exception as e:
            logger.error(f"Failed to load sessions in advanced_modes: {e}")
            
    # Fallback to directory scan if empty
    if not file_paths:
        prefix = f"{session_id}_"
        if os.path.exists(config.UPLOAD_PATH):
            for fname in os.listdir(config.UPLOAD_PATH):
                if fname.startswith(prefix):
                    file_paths.append(fname[len(prefix):])
    return file_paths


def _resolve_selected_docs(candidates: list[str], available_docs: list[str]) -> list[str]:
    """Match candidate doc names against available_docs, tolerant of the "/"
    vs "_" delimiter mismatch between the Files-tab tree paths (sent by the
    frontend as checkbox values, e.g. "Legal AI/NDA (1)/Test_NDA_01.txt")
    and the flat on-disk names /review /compare actually key rows by
    (e.g. "Legal AI_NDA (1)_Test_NDA_01.txt").
    """
    flat_lookup = {d.replace("/", "_").replace("\\", "_"): d for d in available_docs}
    resolved = []
    for c in dict.fromkeys(candidates):
        if c in available_docs:
            resolved.append(c)
        else:
            flat = c.replace("/", "_").replace("\\", "_")
            if flat in flat_lookup:
                resolved.append(flat_lookup[flat])
    return list(dict.fromkeys(resolved))

def _run_review_job(job_id: str, session_id: str, doc_names: list, question: str, store_ref: dict, locks_ref: dict):
    """Background job for review mode with NLP prompt.
    
    Dynamically generates columns based on the prompt, then uses wiki-synthesized content
    instead of raw document text where available, dramatically reducing prompt size and latency.
    """
    lock = locks_ref[job_id]

    def _extract_worker(doc_name, col_name):
        # C7: try metadata cache before firing an LLM call
        cached_res = _lookup_cached_metadata(session_id, doc_name, col_name)
        if cached_res is not None:
            with lock:
                store_ref[job_id]["rows"][doc_name][col_name] = cached_res
                store_ref[job_id]["completed"] += 1
            return

        try:
            res = _extract_with_retrieval(session_id, doc_name, col_name, broad_hint=_is_open_ended_column(col_name))
        except Exception as e:
            logger.error(f"Worker failed for {doc_name}/{col_name}: {e}")
            res = {"value": None, "confidence": 0.0, "quote": None}
        with lock:
            store_ref[job_id]["rows"][doc_name][col_name] = res
            store_ref[job_id]["completed"] += 1
            if res.get("confidence", 0.0) < 0.5 or res.get("value") is None:
                store_ref[job_id]["flagged"].append([doc_name, col_name])

    try:
        # Step 0: Get available documents in the session
        available_docs = _get_session_file_paths(session_id)

        # Step 1: Generate columns, and infer target documents only if the
        # user didn't already pick any — skipping doc inference when docs are
        # already selected keeps this prompt small (the full available_docs
        # list can run into the hundreds) and avoids burning the fast model's
        # completion budget enumerating documents nobody asked it to infer.
        needs_doc_inference = not doc_names
        if needs_doc_inference:
            doc_inference_block = f"""
Available Documents in this session:
{json.dumps(available_docs, indent=1)}
"""
            task_2 = """
2. Infer which of the Available Documents the user wants to review.
   - If the user explicitly mentions document names (or abbreviations/substrings) in their query, map them to the matching document(s) from the Available Documents list.
   - If the user implies certain types of documents or uses keywords (e.g. "all NDA agreements", "the service contract"), select all matching documents from the Available Documents list.
   - If the user does not specify any documents or implies reviewing everything (e.g. "Review these documents"), select ALL Available Documents.
   - Return the inferred documents as a list of exact filenames from the Available Documents list.
"""
            result_shape = '{"columns": ["col1", "col2", ...], "inferred_documents": ["doc_file1", "doc_file2", ...]}'
        else:
            doc_inference_block = ""
            task_2 = ""
            result_shape = '{"columns": ["col1", "col2", ...]}'

        prompt = f"""\
{doc_inference_block}
User Query: {question}

Based on the User Query, perform the following:
1. List the specific factual columns/aspects that need to be extracted from the document(s).
   If the query asks for specific items (e.g., "deliverables, fees, payment terms"), list those exact items as columns.
   If the query is open-ended (e.g., "Summarize this agreement"), list 4-6 key legal or commercial columns to extract.
   Keep column names short (1-5 words).
{task_2}
Return JSON only, no preamble or explanation:
{result_shape}"""

        columns = []
        inferred = []
        try:
            raw, _ = llm.fast_ask(prompt, max_tokens=450)
            import re
            raw = re.sub(r'<reasoning>.*?</reasoning>', '', raw, flags=re.DOTALL)
            raw = re.sub(r'```json', '', raw)
            raw = re.sub(r'```', '', raw)
            parsed = json.loads(raw.strip())
            columns = parsed.get("columns", [])
            inferred = parsed.get("inferred_documents", [])
        except Exception as e:
            logger.error(f"Failed to generate columns/inferred documents: {e}")
            
        if not columns:
            columns = ["Extracted Information"]
            
        # Merge manual selections and inferred documents, resolving them
        # against the available docs (tolerant of "/" vs "_" delimiters)
        all_selected_docs = _resolve_selected_docs(doc_names + inferred, available_docs)

        if not all_selected_docs:
            raise ValueError("No documents were selected or could be inferred from your query.")

        with lock:
            # Re-initialize the rows and columns in the store
            store_ref[job_id]["rows"] = {d: {} for d in all_selected_docs}
            store_ref[job_id]["columns"] = columns
            store_ref[job_id]["total"] = len(all_selected_docs) * len(columns)

        # Step 2: Extract cells, chunked into ≤_MAX_DOCS_PER_RUN-document passes
        # instead of one unbounded batch of concurrent LLM calls. Context is
        # retrieved per (doc, column) inside _extract_worker, scoped by query.
        doc_chunks = _chunk_docs(all_selected_docs)
        if len(doc_chunks) > 1:
            logger.info(
                f"Review job {job_id}: {len(all_selected_docs)} docs exceeds "
                f"{_MAX_DOCS_PER_RUN}-doc cap — running in {len(doc_chunks)} chunks"
            )

        for chunk in doc_chunks:
            with concurrent.futures.ThreadPoolExecutor(max_workers=5) as executor:
                futures = [
                    executor.submit(_extract_worker, doc_name, col_name)
                    for doc_name in chunk
                    for col_name in columns
                ]

                # Wait with timeout to prevent indefinite hanging
                done, not_done = concurrent.futures.wait(futures, timeout=_FUTURE_TIMEOUT * len(futures) / 5 + 30)

                # Cancel any stragglers
                for f in not_done:
                    f.cancel()
                    with lock:
                        store_ref[job_id]["completed"] += 1

        with lock:
            store_ref[job_id]["status"] = "complete"
            
        wiki._log_event(session_id, "REVIEW", f"Job: {job_id} | Docs: {len(all_selected_docs)} | Cols: {len(columns)} | Flagged: {len(store_ref[job_id]['flagged'])}")
            
    except Exception as e:
        logger.error(f"Review job {job_id} failed: {e}")
        with lock:
            store_ref[job_id]["status"] = "error"
            store_ref[job_id]["error"] = str(e)


def _get_scoped_wiki_pages(session_id: str, doc_name: str) -> dict:
    """Filter wiki index to pages belonging to a specific source document.
    
    Pages are matched using two strategies:
    1. Title-based: Wiki page titles follow the convention 'Topic (source_filename)'.
       We check if the parenthesized suffix in the title matches the doc_name.
    2. Field-based: Pages ingested after the source_doc field was added store it directly.
    """
    index = wiki._load_index(session_id)
    pages = index.get("pages", {})
    
    # Build all possible name variants for matching
    flat_name = doc_name.replace("/", "_").replace("\\", "_")
    full_disk_name = f"{session_id}_{flat_name}"
    
    # Normalize for comparison (lowercase, no extension)
    variants = set()
    for name in [doc_name, flat_name, full_disk_name]:
        variants.add(name.lower())
        # Also add without file extension
        base = name.rsplit(".", 1)[0] if "." in name else name
        variants.add(base.lower())
    
    scoped = {}
    for title, page_data in pages.items():
        if not isinstance(page_data, dict):
            continue
        
        matched = False
        
        # Strategy 1: Check source_doc field
        src = page_data.get("source_doc", "")
        if src and src.lower() in variants:
            matched = True
        
        # Strategy 2: Check if the page title contains the doc name in parentheses
        if not matched:
            import re
            paren_match = re.search(r'\(([^)]+)\)\s*$', title)
            if paren_match:
                title_doc = paren_match.group(1).strip()
                title_doc_lower = title_doc.lower()
                if title_doc_lower in variants:
                    matched = True
                # Also check if title_doc ends with any of the base filenames
                if not matched:
                    for v in variants:
                        if title_doc_lower.endswith(v) or v.endswith(title_doc_lower):
                            matched = True
                            break
        
        if matched:
            scoped[title] = page_data
    
    logger.info(f"Scoped wiki lookup for '{doc_name}': found {len(scoped)} pages (out of {len(pages)} total)")
    return scoped

def _run_compare_job(job_id: str, session_id: str, doc_names: list, question: str, 
                     uploaded_text: Optional[str], uploaded_name: Optional[str], temp_path: Optional[str],
                     store_ref: dict, locks_ref: dict):
    """Background job for compare mode.
    
    Optimisations over the original implementation:
    - Uses fast_ask for cell extraction calls
    - Batches outlier detection into a single LLM call instead of per-aspect
    - Uses wiki content with BM25 pre-filtering
    """
    lock = locks_ref[job_id]
    import re
    
    try:
        # STEP 0 - ASPECT IDENTIFICATION AND DOCUMENT INFERENCE
        available_docs = _get_session_file_paths(session_id)

        # Skip document inference (and the potentially hundreds-long available_docs
        # list) when the user already picked documents — see the matching comment
        # in _run_review_job for why this matters for the fast model's budget.
        needs_doc_inference = not doc_names
        if needs_doc_inference:
            doc_inference_block = f"""
Available Documents in this session:
{json.dumps(available_docs, indent=1)}
"""
            task_2 = """
2. Infer which of the Available Documents the user wants to compare.
   - If the user explicitly mentions document names (or abbreviations/substrings) in their query, map them to the matching document(s) from the Available Documents list.
   - If the user implies certain types of documents or uses keywords (e.g. "all NDA agreements", "the service contract"), select all matching documents from the Available Documents list.
   - If the user does not specify any documents or implies comparing everything (e.g. "Compare these documents"), select ALL Available Documents.
   - Return the inferred documents as a list of exact filenames from the Available Documents list.
"""
            result_shape = '{"aspects": ["Aspect 1", "Aspect 2", ...], "inferred_documents": ["doc_file1", "doc_file2", ...]}'
        else:
            doc_inference_block = ""
            task_2 = ""
            result_shape = '{"aspects": ["Aspect 1", "Aspect 2", ...]}'

        aspect_prompt = f"""\
{doc_inference_block}
User Query: {question}

Based on the User Query, perform the following:
1. List the specific factual aspects that need to be extracted for comparison.
   If the query explicitly asks for certain points (e.g., "limits of liability, liability caps"), use those exact points as the aspects.
   If the query is a complex multi-sentence narrative (e.g., "Identify which is most favourable to Tata. Check IP ownership, indemnities, and termination"), extract the core legal/commercial items needed to answer the overall question.
   If the query is open-ended (e.g., "Compare these documents"), list 4-6 key legal and commercial aspects to compare.
   Aspects must be concrete, extractable data points (e.g., "Liability Cap", "Termination Period"), NOT abstract concepts or full sentences.
   Keep aspect names short (1-5 words).
{task_2}
Return JSON only, no preamble or explanation:
{result_shape}"""

        aspects = []
        inferred = []
        try:
            raw, _ = llm.fast_ask(aspect_prompt, max_tokens=450)
            raw = re.sub(r'<reasoning>.*?</reasoning>', '', raw, flags=re.DOTALL)
            raw = re.sub(r'```json', '', raw)
            raw = re.sub(r'```', '', raw)
            parsed = json.loads(raw.strip())
            aspects = parsed.get("aspects", [])
            inferred = parsed.get("inferred_documents", [])
        except Exception as e:
            logger.error(f"Failed to generate aspects and inferred documents: {e}")
            
        if not aspects:
            aspects = ["Comparison Details"]
            
        # Merge manually selected and inferred documents, resolving them
        # against the available docs (tolerant of "/" vs "_" delimiters)
        all_selected_docs = _resolve_selected_docs(doc_names + inferred, available_docs)

        if not all_selected_docs and not uploaded_name:
            raise ValueError("No documents were selected or could be inferred from your query.")

        # STEP 1 - NORMALIZE SOURCES
        # Sources carry only identity here — per-(source, aspect) text is now
        # fetched at extraction time via document-scoped hybrid retrieval
        # (_extract_with_retrieval), not pre-concatenated once for every aspect.
        with lock:
            store_ref[job_id]["stage"] = "retrieving"

        doc_sources = [{"name": doc_name, "type": "wiki"} for doc_name in all_selected_docs]

        upload_source = None
        if uploaded_text and uploaded_name:
            upload_source = {"name": uploaded_name, "type": "upload", "text": uploaded_text[:12000], "label": f"{uploaded_name} ⚡"}

        all_sources = doc_sources + ([upload_source] if upload_source else [])
        with lock:
            store_ref[job_id]["sources"] = [{"name": s["name"], "type": s["type"], "label": s.get("label", s["name"])} for s in all_sources]

        # STEP 2 - INITIALIZE ASPECTS AND TABLE IN STORE
        with lock:
            store_ref[job_id]["stage"] = "extracting"
            store_ref[job_id]["aspects"] = aspects
            for aspect in aspects:
                store_ref[job_id]["table"][aspect] = {}

        # STEP 3 - EXTRACT PER SOURCE PER ASPECT (parallel with fast_ask),
        # chunked into ≤_MAX_DOCS_PER_RUN-document passes instead of one
        # unbounded batch of concurrent LLM calls.
        def _extract_compare_worker(source_dict, aspect):
            doc_key = source_dict.get("label", source_dict["name"])
            if source_dict["type"] == "upload":
                try:
                    res = extract_cell(source_dict["text"], aspect)
                except Exception as e:
                    logger.error(f"Compare extract failed {source_dict['name']}/{aspect}: {e}")
                    res = {"value": None, "confidence": 0.0, "quote": None}
            else:
                # C7: try metadata cache before firing an LLM call
                res = _lookup_cached_metadata(session_id, source_dict["name"], aspect)
                if res is None:
                    try:
                        res = _extract_with_retrieval(
                            session_id, source_dict["name"], aspect,
                            broad_hint=_is_open_ended_column(aspect),
                        )
                    except Exception as e:
                        logger.error(f"Compare extract failed {source_dict['name']}/{aspect}: {e}")
                        res = {"value": None, "confidence": 0.0, "quote": None}
            with lock:
                store_ref[job_id]["table"][aspect][doc_key] = res

        doc_chunks = _chunk_docs(all_selected_docs) or [[]]  # keep one (possibly empty) chunk so an upload-only compare still runs
        if len(doc_chunks) > 1:
            logger.info(
                f"Compare job {job_id}: {len(all_selected_docs)} docs exceeds "
                f"{_MAX_DOCS_PER_RUN}-doc cap — running in {len(doc_chunks)} chunks"
            )
        doc_source_by_name = {s["name"]: s for s in doc_sources}

        for i, chunk in enumerate(doc_chunks):
            chunk_sources = [doc_source_by_name[d] for d in chunk]
            if upload_source and i == 0:
                chunk_sources = chunk_sources + [upload_source]

            with concurrent.futures.ThreadPoolExecutor(max_workers=5) as executor:
                futures = [
                    executor.submit(_extract_compare_worker, s, a)
                    for s in chunk_sources
                    for a in aspects
                ]

                # Wait with timeout
                total_cells = len(futures)
                timeout = max(60, total_cells * 15)  # ~15s per cell, minimum 60s
                done, not_done = concurrent.futures.wait(futures, timeout=timeout)

                # Fill in failures for timed-out cells
                for f in not_done:
                    f.cancel()


        # STEP 4 - OUTLIER DETECTION (batched into a single LLM call)
        with lock:
            store_ref[job_id]["stage"] = "analyzing"
            
        # Build a compact summary of all values for batch outlier detection
        all_values = {}
        for aspect in aspects:
            aspect_vals = {}
            for s in all_sources:
                doc_key = s.get("label", s["name"])
                cell_data = store_ref[job_id]["table"].get(aspect, {}).get(doc_key, {})
                val = cell_data.get("value")
                if val:
                    aspect_vals[doc_key] = val
            if aspect_vals:
                all_values[aspect] = aspect_vals
 
        outliers = []
        if all_values:
            outlier_prompt = f"""\
Compare the extracted values below across documents. Identify ONLY contradictions
or significant differences that are DIRECTLY visible in the data provided.
DO NOT infer, speculate, or flag differences that are not clearly present.
Return JSON only — an array of objects:
[{{"aspect": str, "doc": str, "reason": str}}] or []
Return [] if no clear contradictions exist.

Extracted values:
{json.dumps(all_values, indent=1)}"""

            try:
                raw, _ = llm.fast_ask(outlier_prompt, max_tokens=1000)
                raw = re.sub(r'<reasoning>.*?</reasoning>', '', raw, flags=re.DOTALL)
                raw = re.sub(r'```json', '', raw)
                raw = re.sub(r'```', '', raw)
                parsed_outliers = json.loads(raw.strip())
                if isinstance(parsed_outliers, list):
                    outliers = parsed_outliers
            except Exception as e:
                logger.error(f"Outlier detection failed: {e}")
                
        with lock:
            store_ref[job_id]["outliers"] = outliers
            
        # STEP 5 - NARRATIVE GENERATION (single call). This is summarizing an
        # already-extracted table, not doing fresh legal reasoning over raw
        # text, so the fast/cheap model is sufficient — same tier used for
        # every other Review/Compare call.
        narrative_prompt = f"""\
Question: {question}
Comparison table: {json.dumps(store_ref[job_id]["table"])}
Outliers: {json.dumps(outliers)}

Write a well-structured legal synthesis based ONLY on the comparison table data above.

STRICT RULES:
- Format your response using markdown with clear headings, bullet points, and bold text for readability.
- Write fluid, cohesive summary paragraphs describing trends across the documents. DO NOT output repetitive nested lists detailing every single document's individual value. 
- Group similar findings together into single sentences (e.g. "Most documents do not specify a notice period, except [1] which requires 30 days").
- Do NOT redundantly enumerate "not specified" for every document. Synthesize it.
- Cite your sources using standard IEEE format inline (e.g., [1], [2], [3]). Do NOT mention the exact document names anywhere in the paragraph text.
- ONLY state facts that appear in the comparison table. DO NOT add information not present in the data.
- Flag contradictions explicitly, citing both documents and their conflicting values verbatim.
- DO NOT invent legal conclusions, implications, or recommendations beyond what the data shows.
- PROPER CITATIONS (CRITICAL): You MUST create a "References" list at the very end of your answer starting with a "References" heading. Each entry must strictly follow this pattern: "[X] File_Name.pdf, Clause/Page | Quote: <exact verbatim quote from the text>" (e.g. "[1] Service Agreement 1_redacted.pdf, Clause 14.1 | Quote: The Supplier shall deliver..."). If the exact clause/page or quote is not in the table, just map it as: "[1] Service Agreement 1_redacted.pdf | Quote: <verbatim quote>" or "[1] Service Agreement 1_redacted.pdf". Do not wrap file names in formatting."""

        narrative, _ = llm.fast_ask(narrative_prompt, max_tokens=1500)
        
        # STEP 6 - STORE + COMPLETE
        with lock:
            store_ref[job_id]["narrative"] = narrative
            store_ref[job_id]["status"] = "complete"
            
        wiki._log_event(session_id, "COMPARE", f"Job: {job_id} | Docs: {len(all_selected_docs) + (1 if uploaded_text else 0)} | Aspects: {len(aspects)} | Outliers: {len(outliers)}")
        
    except Exception as e:
        logger.error(f"Compare job {job_id} failed: {e}")
        with lock:
            store_ref[job_id]["status"] = "error"
            store_ref[job_id]["error"] = str(e)
            
    finally:
        if temp_path and os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except Exception as e:
                logger.error(f"Failed to clean up temp file {temp_path}: {e}")
